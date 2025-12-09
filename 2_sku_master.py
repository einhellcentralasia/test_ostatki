#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import io
import os
import sys
import time
import json
from dataclasses import dataclass
from typing import Iterable, Optional, Tuple, List

import pandas as pd
import requests
from openpyxl import load_workbook
from urllib.parse import quote


# =========================
# CONFIG (edit here if needed)
# =========================
SITE_HOSTNAME = os.getenv("SP_SITE_HOSTNAME", "bavatools.sharepoint.com")
SITE_PATH     = os.getenv("SKU_SITE_PATH", "/sites/Einhell_common")
FILE_PATH     = os.getenv("SKU_XLSX_PATH", "/Shared Documents/General/_system_files/Bava_data.xlsx")
TABLE_NAME    = os.getenv("SKU_TABLE_NAME", "list_stock_final_table_transit_table")

OUTPUT_NAME   = os.getenv("SKU_OUTPUT_NAME", "SKU")  # /data/<OUTPUT_NAME>/<OUTPUT_NAME>.(parquet|csv|json)
KEEP_COLUMNS  = ["SKU", "Model", "RSP", "ETA_Almaty"]


# -------------------------
# Poka-yoke / error handling
# -------------------------

def die(msg: str, code: int = 2) -> None:
    print(f"❌ {msg}", file=sys.stderr)
    raise SystemExit(code)

def env(name: str, default: Optional[str] = None, required: bool = True) -> str:
    val = os.getenv(name, default)
    if required and (val is None or str(val).strip() == ""):
        die(f"Missing required env var: {name}")
    return str(val)

def now_ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


# -------------------------
# Microsoft Graph helpers
# -------------------------

GRAPH = "https://graph.microsoft.com/v1.0"

@dataclass
class GraphCtx:
    token: str
    session: requests.Session

def new_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"Accept": "application/json"})
    return s

def request_raw(ctx: GraphCtx, method: str, url: str, *, params=None, timeout: int = 60) -> requests.Response:
    max_tries = 8
    backoff = 1.0
    headers = {"Authorization": f"Bearer {ctx.token}", "Accept": "application/json"}

    for attempt in range(1, max_tries + 1):
        resp = ctx.session.request(method, url, headers=headers, params=params, timeout=timeout)

        if resp.status_code in (429, 500, 502, 503, 504):
            retry_after = resp.headers.get("Retry-After")
            try:
                sleep_s = float(retry_after) if retry_after else backoff
            except ValueError:
                sleep_s = backoff
            print(f"⚠️ {now_ts()} Graph {resp.status_code} (attempt {attempt}/{max_tries}), sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)
            backoff = min(backoff * 1.8, 20.0)
            continue

        return resp

    die(f"Graph failed after retries on {url}")
    return resp  # unreachable

def request_json_ok(ctx: GraphCtx, method: str, url: str, *, params=None, expected=(200,)) -> dict:
    resp = request_raw(ctx, method, url, params=params, timeout=60)
    if resp.status_code not in expected:
        die(f"Graph error {resp.status_code} on {url}\nResponse: {resp.text[:2000]}")
    return resp.json()

def request_bytes(ctx: GraphCtx, url: str) -> bytes:
    max_tries = 8
    backoff = 1.0
    headers = {"Authorization": f"Bearer {ctx.token}"}

    for attempt in range(1, max_tries + 1):
        resp = ctx.session.get(url, headers=headers, stream=True, timeout=120)
        if resp.status_code == 200:
            return resp.content

        if resp.status_code in (429, 500, 502, 503, 504):
            retry_after = resp.headers.get("Retry-After")
            try:
                sleep_s = float(retry_after) if retry_after else backoff
            except ValueError:
                sleep_s = backoff
            print(f"⚠️ {now_ts()} Download {resp.status_code} (attempt {attempt}/{max_tries}), sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)
            backoff = min(backoff * 1.8, 20.0)
            continue

        die(f"Download error {resp.status_code} for {url}: {resp.text[:1000]}")

    die(f"Download failed after retries: {url}")
    return b""

def get_app_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default",
    }
    resp = requests.post(token_url, data=data, timeout=60)
    if resp.status_code != 200:
        die(f"Token request failed {resp.status_code}: {resp.text[:2000]}")
    tok = resp.json().get("access_token")
    if not tok:
        die("Token response missing access_token")
    return tok

def normalize_sp_path(p: str) -> str:
    p2 = p.strip().replace("\\", "/").lstrip("/")
    while "//" in p2:
        p2 = p2.replace("//", "/")
    low = p2.lower()
    for prefix in ("shared documents/", "documents/"):
        if low.startswith(prefix):
            p2 = p2[len(prefix):]
            break
    return p2

def graph_get_site_id(ctx: GraphCtx, hostname: str, site_path: str) -> str:
    url = f"{GRAPH}/sites/{hostname}:{site_path}"
    js = request_json_ok(ctx, "GET", url, expected=(200,))
    site_id = js.get("id")
    if not site_id:
        die("Could not resolve site id. Check hostname/site_path.")
    return site_id

def graph_get_drive_id(ctx: GraphCtx, site_id: str) -> str:
    url = f"{GRAPH}/sites/{site_id}/drive"
    js = request_json_ok(ctx, "GET", url, expected=(200,))
    drive_id = js.get("id")
    if not drive_id:
        die("Could not resolve drive id for the site.")
    return drive_id

def try_get_item_id_by_path(ctx: GraphCtx, drive_id: str, path: str) -> Optional[str]:
    path_clean = normalize_sp_path(path)
    path_enc = quote(path_clean, safe="/")
    url = f"{GRAPH}/drives/{drive_id}/root:/{path_enc}"
    resp = request_raw(ctx, "GET", url, timeout=60)
    if resp.status_code == 200:
        return resp.json().get("id")
    if resp.status_code == 404:
        return None
    die(f"Unexpected status {resp.status_code} resolving path.\nURL: {url}\nBody: {resp.text[:2000]}")
    return None


# -------------------------
# Excel table extraction (named table)
# -------------------------

def iter_ws_tables(ws):
    t = getattr(ws, "tables", None)
    if not t:
        return
    try:
        for name in list(t.keys()):
            yield name, t[name]
    except Exception:
        try:
            for tbl in t:
                yield getattr(tbl, "name", None), tbl
        except Exception:
            return

def extract_table_df(ws, ref: str) -> pd.DataFrame:
    cells = ws[ref]
    rows = [[c.value for c in row] for row in cells]
    if not rows:
        return pd.DataFrame()
    headers = [("" if x is None else str(x).strip()) for x in rows[0]]
    data = rows[1:]
    return pd.DataFrame(data, columns=headers)

def read_named_table_from_xlsx_bytes(xlsx_bytes: bytes, table_name: str) -> Tuple[pd.DataFrame, str]:
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    target = table_name.strip().lower()

    found_names: List[str] = []
    for ws in wb.worksheets:
        for tname, tbl in iter_ws_tables(ws) or []:
            if tname:
                found_names.append(tname)
            if tname and tname.strip().lower() == target:
                ref = getattr(tbl, "ref", None)
                if not ref:
                    die(f"Table '{table_name}' found but missing ref range")
                df = extract_table_df(ws, ref)
                return df, "OK(table)"

    die(
        f"Named table '{table_name}' not found in workbook.\n"
        f"Tables found: {found_names if found_names else '[none]'}"
    )
    return pd.DataFrame(), "unreachable"


# -------------------------
# Output helpers
# -------------------------

def canonical(s: str) -> str:
    return "".join(str(s).strip().lower().split())

def select_columns_case_insensitive(df: pd.DataFrame, wanted: List[str]) -> pd.DataFrame:
    col_map = {canonical(c): c for c in df.columns}
    missing = []
    picked = {}
    for w in wanted:
        key = canonical(w)
        if key in col_map:
            picked[w] = col_map[key]
        else:
            missing.append(w)

    if missing:
        die(
            f"Missing required columns in table: {missing}\n"
            f"Available columns: {list(df.columns)}"
        )

    out = df[[picked[w] for w in wanted]].copy()
    out.columns = wanted
    return out

def coerce_eta_date(df: pd.DataFrame) -> pd.DataFrame:
    # ETA_Almaty is dd.mm.yyyy
    df["ETA_Almaty"] = pd.to_datetime(df["ETA_Almaty"], dayfirst=True, errors="coerce").dt.date
    return df

def save_outputs(df: pd.DataFrame, name: str) -> None:
    out_dir = os.path.join("data", name)
    ensure_dir(out_dir)

    df.to_parquet(os.path.join(out_dir, f"{name}.parquet"), index=False)

    df.to_csv(
        os.path.join(out_dir, f"{name}.csv"),
        index=False,
        encoding="utf-8",
        date_format="%d.%m.%Y",
    )

    with open(os.path.join(out_dir, f"{name}.json"), "w", encoding="utf-8") as f:
        json.dump(df.to_dict(orient="records"), f, ensure_ascii=False, indent=2)


# -------------------------
# Main
# -------------------------

def main() -> None:
    t0 = time.perf_counter()

    tenant_id = env("TENANT_ID")
    client_id = env("CLIENT_ID")
    client_secret = env("CLIENT_SECRET")

    print(f"🟢 {now_ts()} Start SKU master sync")
    print(f"   Site: https://{SITE_HOSTNAME}{SITE_PATH}")
    print(f"   File: {FILE_PATH}")
    print(f"   Table: {TABLE_NAME}")
    print(f"   Output: data/{OUTPUT_NAME}/")

    token = get_app_token(tenant_id, client_id, client_secret)
    ctx = GraphCtx(token=token, session=new_session())

    site_id = graph_get_site_id(ctx, SITE_HOSTNAME, SITE_PATH)
    drive_id = graph_get_drive_id(ctx, site_id)

    item_id = try_get_item_id_by_path(ctx, drive_id, FILE_PATH)
    if not item_id:
        die(f"File not found at path: {FILE_PATH}")

    xlsx_bytes = request_bytes(ctx, f"{GRAPH}/drives/{drive_id}/items/{item_id}/content")
    df_raw, _ = read_named_table_from_xlsx_bytes(xlsx_bytes, TABLE_NAME)

    if df_raw.empty:
        die(f"Table '{TABLE_NAME}' extracted but is empty.")

    df = select_columns_case_insensitive(df_raw, KEEP_COLUMNS)
    df = coerce_eta_date(df)

    df["SKU"] = df["SKU"].astype(str).str.strip()
    df["Model"] = df["Model"].astype(str).str.strip()

    save_outputs(df, OUTPUT_NAME)

    total_s = round(time.perf_counter() - t0, 2)
    print(f"✅ Done. Rows={len(df)} | Runtime={total_s}s")
    print(f"📁 Outputs: /data/{OUTPUT_NAME}/{OUTPUT_NAME}.parquet|csv|json")

if __name__ == "__main__":
    main()
