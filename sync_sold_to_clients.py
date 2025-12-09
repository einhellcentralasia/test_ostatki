#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import io
import os
import sys
import time
import json
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests
import yaml
from openpyxl import load_workbook
from urllib.parse import quote


# -------------------------
# Poka-yoke / error handling
# -------------------------

def die(msg: str, code: int = 2) -> None:
    print(f"❌ {msg}", file=sys.stderr)
    raise SystemExit(code)

def env(name: str, default: Optional[str] = None, required: bool = True) -> str:
    val = os.getenv(name, default)
    if required and (val is None or str(val).strip() == ""):
        die(f"Missing required env var: {name}")
    return str(val)

def now_ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


# -------------------------
# Microsoft Graph helpers
# -------------------------

GRAPH = "https://graph.microsoft.com/v1.0"

@dataclass
class GraphCtx:
    token: str
    session: requests.Session

def new_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"Accept": "application/json"})
    return s

def request_raw(ctx: GraphCtx, method: str, url: str, *, params=None, timeout: int = 60) -> requests.Response:
    max_tries = 8
    backoff = 1.0
    headers = {"Authorization": f"Bearer {ctx.token}", "Accept": "application/json"}

    for attempt in range(1, max_tries + 1):
        resp = ctx.session.request(method, url, headers=headers, params=params, timeout=timeout)

        if resp.status_code in (429, 500, 502, 503, 504):
            retry_after = resp.headers.get("Retry-After")
            try:
                sleep_s = float(retry_after) if retry_after else backoff
            except ValueError:
                sleep_s = backoff
            print(f"⚠️ {now_ts()} Graph {resp.status_code} (attempt {attempt}/{max_tries}), sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)
            backoff = min(backoff * 1.8, 20.0)
            continue

        return resp

    die(f"Graph failed after retries on {url}")
    return resp  # unreachable

def request_json_ok(ctx: GraphCtx, method: str, url: str, *, params=None, expected=(200,)) -> dict:
    resp = request_raw(ctx, method, url, params=params, timeout=60)
    if resp.status_code not in expected:
        die(f"Graph error {resp.status_code} on {url}\nResponse: {resp.text[:2000]}")
    return resp.json()

def request_bytes(ctx: GraphCtx, url: str) -> bytes:
    max_tries = 8
    backoff = 1.0
    headers = {"Authorization": f"Bearer {ctx.token}"}

    for attempt in range(1, max_tries + 1):
        resp = ctx.session.get(url, headers=headers, stream=True, timeout=120)
        if resp.status_code == 200:
            return resp.content

        if resp.status_code in (429, 500, 502, 503, 504):
            retry_after = resp.headers.get("Retry-After")
            try:
                sleep_s = float(retry_after) if retry_after else backoff
            except ValueError:
                sleep_s = backoff
            print(f"⚠️ {now_ts()} Download {resp.status_code} (attempt {attempt}/{max_tries}), sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)
            backoff = min(backoff * 1.8, 20.0)
            continue

        die(f"Download error {resp.status_code} for {url}: {resp.text[:1000]}")

    die(f"Download failed after retries: {url}")
    return b""

def get_app_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default",
    }
    resp = requests.post(token_url, data=data, timeout=60)
    if resp.status_code != 200:
        die(f"Token request failed {resp.status_code}: {resp.text[:2000]}")
    tok = resp.json().get("access_token")
    if not tok:
        die("Token response missing access_token")
    return tok

def normalize_sp_path(p: str) -> str:
    p2 = p.strip().replace("\\", "/").lstrip("/")
    while "//" in p2:
        p2 = p2.replace("//", "/")
    low = p2.lower()
    for prefix in ("shared documents/", "documents/"):
        if low.startswith(prefix):
            p2 = p2[len(prefix):]
            break
    return p2

def graph_get_site_id(ctx: GraphCtx, hostname: str, site_path: str) -> str:
    url = f"{GRAPH}/sites/{hostname}:{site_path}"
    js = request_json_ok(ctx, "GET", url, expected=(200,))
    site_id = js.get("id")
    if not site_id:
        die("Could not resolve site id. Check site_hostname/site_path in YAML.")
    return site_id

def graph_get_drive_id(ctx: GraphCtx, site_id: str) -> str:
    url = f"{GRAPH}/sites/{site_id}/drive"
    js = request_json_ok(ctx, "GET", url, expected=(200,))
    drive_id = js.get("id")
    if not drive_id:
        die("Could not resolve drive id for the site.")
    return drive_id

def try_get_item_id_by_path(ctx: GraphCtx, drive_id: str, path: str) -> Optional[str]:
    path_clean = normalize_sp_path(path)
    path_enc = quote(path_clean, safe="/")
    url = f"{GRAPH}/drives/{drive_id}/root:/{path_enc}"
    resp = request_raw(ctx, "GET", url, timeout=60)
    if resp.status_code == 200:
        return resp.json().get("id")
    if resp.status_code == 404:
        return None
    die(f"Unexpected status {resp.status_code} resolving folder path.\nURL: {url}\nBody: {resp.text[:2000]}")
    return None

def graph_list_children(ctx: GraphCtx, drive_id: str, folder_item_id: str) -> Iterable[dict]:
    url = f"{GRAPH}/drives/{drive_id}/items/{folder_item_id}/children"
    params = {"$top": "200"}
    while True:
        js = request_json_ok(ctx, "GET", url, params=params, expected=(200,))
        for it in js.get("value", []):
            yield it
        nxt = js.get("@odata.nextLink")
        if not nxt:
            break
        url = nxt
        params = None

def graph_walk_files(ctx: GraphCtx, drive_id: str, folder_item_id: str, recursive: bool) -> List[dict]:
    out = []
    stack = [folder_item_id]
    while stack:
        fid = stack.pop()
        for it in graph_list_children(ctx, drive_id, fid):
            name = it.get("name", "")
            if name.startswith("~$"):
                continue
            is_folder = "folder" in it
            is_file = "file" in it

            if is_folder and recursive:
                cid = it.get("id")
                if cid:
                    stack.append(cid)
                continue
            if is_file:
                out.append(it)
    return out


# -------------------------
# Excel parsing helpers (SKU+Qty only)
# -------------------------

def norm(s: str) -> str:
    return "".join(str(s).strip().lower().split())

SKU_HEADERS = {"sku", "артикул", "арт", "item", "code"}
QTY_HEADERS = {"qty", "qt", "quantity", "кол-во", "количество", "count", "pcs"}

def find_col_indices(headers: List[str]) -> Tuple[Optional[int], Optional[int]]:
    sku_idx = None
    qty_idx = None
    for i, h in enumerate(headers):
        h2 = norm(h)
        if sku_idx is None and h2 in SKU_HEADERS:
            sku_idx = i
        if qty_idx is None and h2 in QTY_HEADERS:
            qty_idx = i
    return sku_idx, qty_idx

def iter_ws_tables(ws):
    t = getattr(ws, "tables", None)
    if not t:
        return
    try:
        for name in list(t.keys()):
            yield name, t[name]
    except Exception:
        try:
            for tbl in t:
                yield getattr(tbl, "name", None), tbl
        except Exception:
            return

def extract_df_from_range(ws, ref: str) -> Optional[pd.DataFrame]:
    cells = ws[ref]
    rows = [[c.value for c in row] for row in cells]
    if not rows or len(rows) < 2:
        return pd.DataFrame(columns=["SKU", "Qty"])
    headers = [str(x) if x is not None else "" for x in rows[0]]
    sku_idx, qty_idx = find_col_indices(headers)
    if sku_idx is None or qty_idx is None:
        return None
    skus, qtys = [], []
    for r in rows[1:]:
        if sku_idx >= len(r) or qty_idx >= len(r):
            continue
        sku = r[sku_idx]
        qty = r[qty_idx]
        if sku is None or str(sku).strip() == "":
            continue
        try:
            q = float(qty) if qty is not None and str(qty).strip() != "" else 0.0
        except Exception:
            q = 0.0

        # ✅ CHANGE: ignore negative numbers
        if q < 0:
            continue

        skus.append(str(sku).strip())
        qtys.append(q)
    return pd.DataFrame({"SKU": skus, "Qty": qtys})

def fallback_scan_headers(ws, max_rows: int = 5000, max_cols: int = 80) -> Optional[pd.DataFrame]:
    mr = min(ws.max_row or 0, max_rows)
    mc = min(ws.max_column or 0, max_cols)
    if mr <= 0 or mc <= 0:
        return None

    header_row = None
    sku_col = None
    qty_col = None

    for r in range(1, mr + 1):
        vals = []
        for c in range(1, mc + 1):
            v = ws.cell(row=r, column=c).value
            vals.append("" if v is None else str(v))
        sku_idx, qty_idx = find_col_indices(vals)
        if sku_idx is not None and qty_idx is not None:
            header_row = r
            sku_col = sku_idx + 1
            qty_col = qty_idx + 1
            break

    if header_row is None:
        return None

    skus, qtys = [], []
    blank_streak = 0
    for r in range(header_row + 1, mr + 1):
        sku = ws.cell(row=r, column=sku_col).value
        qty = ws.cell(row=r, column=qty_col).value
        is_blank = (sku is None or str(sku).strip() == "") and (qty is None or str(qty).strip() == "")
        if is_blank:
            blank_streak += 1
            if blank_streak >= 3:
                break
            continue
        blank_streak = 0
        if sku is None or str(sku).strip() == "":
            continue
        try:
            q = float(qty) if qty is not None and str(qty).strip() != "" else 0.0
        except Exception:
            q = 0.0

        # ✅ CHANGE: ignore negative numbers
        if q < 0:
            continue

        skus.append(str(sku).strip())
        qtys.append(q)

    return pd.DataFrame({"SKU": skus, "Qty": qtys})

def read_sku_qty_from_xlsx_bytes(xlsx_bytes: bytes, table_name: str) -> Tuple[Optional[pd.DataFrame], str]:
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    target = table_name.strip().lower()

    for ws in wb.worksheets:
        for tname, tbl in iter_ws_tables(ws) or []:
            if tname and tname.strip().lower() == target:
                ref = getattr(tbl, "ref", None)
                if not ref:
                    return None, "Table found but missing ref"
                df = extract_df_from_range(ws, ref)
                if df is None:
                    return None, "Table found but SKU/Qty headers not detected"
                return df, "OK(table)"

    for ws in wb.worksheets:
        df2 = fallback_scan_headers(ws)
        if df2 is not None:
            return df2, "OK(fallback)"

    return None, "Table not found + fallback scan failed"


# -------------------------
# Config + pipeline
# -------------------------

def load_config(path: str) -> dict:
    if not os.path.exists(path):
        die(f"Config file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    if not isinstance(cfg, dict):
        die("Config must be a YAML mapping (dict).")
    return cfg

def validate_config(cfg: dict) -> None:
    if not cfg.get("site_hostname"):
        die("Config missing 'site_hostname'")
    if not cfg.get("table_name"):
        die("Config missing 'table_name'")
    sources = cfg.get("sources")
    if not isinstance(sources, list) or not sources:
        die("Config missing 'sources' list or it's empty")

    seen = set()
    for s in sources:
        if not isinstance(s, dict):
            die("Each source must be a dict.")
        name = s.get("name")
        if not name or not isinstance(name, str):
            die("Each source must have a string 'name'")
        if name in seen:
            die(f"Duplicate source name: {name}")
        seen.add(name)
        if not s.get("site_path"):
            die(f"Source '{name}' missing site_path")
        if not s.get("xlsx_path"):
            die(f"Source '{name}' missing xlsx_path")

def run_source(ctx: GraphCtx, site_hostname: str, table_name: str, source: dict) -> Tuple[pd.DataFrame, dict]:
    name = source["name"]
    site_path = source["site_path"]
    xlsx_path = source["xlsx_path"]
    recursive = bool(source.get("recursive", False))

    t0 = time.perf_counter()

    try:
        site_id = graph_get_site_id(ctx, site_hostname, site_path)
        drive_id = graph_get_drive_id(ctx, site_id)
        folder_item_id = try_get_item_id_by_path(ctx, drive_id, xlsx_path)
        if not folder_item_id:
            return pd.DataFrame(columns=["SKU", "Qty"]), {
                "name": name, "status": "error", "reason": "Folder not found",
                "site_path": site_path, "xlsx_path": xlsx_path, "runtime_s": round(time.perf_counter() - t0, 2)
            }

        items = graph_walk_files(ctx, drive_id, folder_item_id, recursive=recursive)
        xlsx_items = [it for it in items if it.get("name", "").lower().endswith(".xlsx")]

        agg: Dict[str, float] = {}
        processed = 0
        skipped = 0
        used_fallback = 0

        for it in xlsx_items:
            item_id = it.get("id")
            if not item_id:
                skipped += 1
                continue

            b = request_bytes(ctx, f"{GRAPH}/drives/{drive_id}/items/{item_id}/content")
            df, reason = read_sku_qty_from_xlsx_bytes(b, table_name)
            if df is None:
                skipped += 1
                continue

            if reason.endswith("(fallback)"):
                used_fallback += 1

            if not df.empty:
                for sku, qty in zip(df["SKU"].tolist(), df["Qty"].tolist()):
                    # qty is already non-negative from extract_df_from_range / fallback_scan_headers
                    agg[sku] = agg.get(sku, 0.0) + float(qty)

            processed += 1

        if agg:
            out = pd.DataFrame([{"SKU": sku, "Qty": agg[sku]} for sku in agg.keys()])
            out = out.sort_values(["SKU"], kind="stable")
            out["Qty"] = out["Qty"].apply(lambda x: int(x) if float(x).is_integer() else float(x))
        else:
            out = pd.DataFrame(columns=["SKU", "Qty"])

        meta = {
            "name": name,
            "status": "ok",
            "site_path": site_path,
            "xlsx_path": xlsx_path,
            "recursive": recursive,
            "files_total": len(items),
            "files_xlsx": len(xlsx_items),
            "processed_xlsx": processed,
            "skipped_xlsx": skipped,
            "used_fallback": used_fallback,
            "unique_skus": int(len(out)),
            "runtime_s": round(time.perf_counter() - t0, 2),
        }
        return out, meta

    except Exception as e:
        return pd.DataFrame(columns=["SKU", "Qty"]), {
            "name": name, "status": "error", "reason": f"Exception: {e}",
            "site_path": site_path, "xlsx_path": xlsx_path, "runtime_s": round(time.perf_counter() - t0, 2)
        }

def save_outputs(df: pd.DataFrame, name: str) -> None:
    out_dir = os.path.join("data", name)
    ensure_dir(out_dir)
    df.to_parquet(os.path.join(out_dir, f"{name}.parquet"), index=False)
    df.to_csv(os.path.join(out_dir, f"{name}.csv"), index=False, encoding="utf-8")

def save_manifest(summaries: List[dict]) -> None:
    ensure_dir("data")
    mf = pd.DataFrame(summaries)
    cols = [
        "name", "status", "reason",
        "site_path", "xlsx_path", "recursive",
        "files_total", "files_xlsx", "processed_xlsx", "skipped_xlsx", "used_fallback",
        "unique_skus", "runtime_s"
    ]
    for c in cols:
        if c not in mf.columns:
            mf[c] = None
    mf = mf[cols]
    mf.to_csv("data/_manifest.csv", index=False, encoding="utf-8")
    with open("data/_manifest.json", "w", encoding="utf-8") as f:
        json.dump(summaries, f, ensure_ascii=False, indent=2)

def main() -> None:
    t_all = time.perf_counter()

    tenant_id = env("TENANT_ID")
    client_id = env("CLIENT_ID")
    client_secret = env("CLIENT_SECRET")
    config_path = env("CONFIG_PATH", default="sharepoint_sources.yml", required=False)

    cfg = load_config(config_path)
    validate_config(cfg)

    site_hostname = cfg["site_hostname"]
    table_name = cfg["table_name"]

    print(f"🟢 {now_ts()} Start. Config={config_path}")
    print(f"   Hostname={site_hostname} | Table={table_name}")

    token = get_app_token(tenant_id, client_id, client_secret)
    ctx = GraphCtx(token=token, session=new_session())

    summaries: List[dict] = []

    for src in cfg["sources"]:
        if not src.get("enabled", True):
            continue

        name = src["name"]
        print(f"\n▶️ Source: {name}")

        df, meta = run_source(ctx, site_hostname, table_name, src)

        if meta.get("status") == "ok":
            save_outputs(df, name)
            print(f"✅ {name}: xlsx={meta.get('files_xlsx')} processed={meta.get('processed_xlsx')} skipped={meta.get('skipped_xlsx')} skus={meta.get('unique_skus')} time={meta.get('runtime_s')}s")
        else:
            save_outputs(pd.DataFrame(columns=["SKU", "Qty"]), name)
            print(f"❌ {name}: {meta.get('reason')} (empty outputs written)")

        summaries.append(meta)

    save_manifest(summaries)

    total_s = round(time.perf_counter() - t_all, 2)
    print("\n📌 Summary:")
    for m in summaries:
        print(f" - {m.get('name')}: {m.get('status')} | skus={m.get('unique_skus', '?')} | time={m.get('runtime_s', '?')}s")

    print(f"\n⏱️ Total runtime: {total_s}s")
    print("📁 Outputs live under /data/<name>/ and manifest under /data/_manifest.(csv|json)")

if __name__ == "__main__":
    main()
